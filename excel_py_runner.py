"""
Python function runner.
Reads and executes sequence of python fuctions, with optional args
from specially formatted spreadsheet.
Functions to be run are defined in actions.py.
"""
import time
import os
import openpyxl
import actions


class Excel(object):
    """Read specially formatted spreadsheet,
        Execute actions specified within it,
        Save new copy of spreadsheet with results of the actions recorded.
        Depends on actions.py for definitions of available actions.
    """
    def __init__(self, filename):
        """Read specially formatted spreadsheet,
        Args:
            filename - spreadsheet fileaname
        """
        self.filename = filename
        self.wb = openpyxl.load_workbook(filename=filename)
        self.fills = {}
        self.define_styles()
        
    def define_styles(self):
        """Define some background colours and store in self.fills"""
        colours = [("red", "FFFF3333"),
                   ("orange", "FFFF8000"),
                   ("yellow", "50FFFF00"),
                   ("green", "FFB2FF66"),
                   ("purple", "FFCC00CC")
                   ]
        for col in colours:
            self.fills[col[0]] = openpyxl.styles.PatternFill(
                       start_color=col[1],
                       end_color=col[1],
                       fill_type='solid')

    def run_tab(self, tabname="Activities"):
        """Process sequence of actions from the specified spreadsheet tab
        Write results to new spreadsheet in results sub-folder with date/time
        in filename.
        Args:
            tabname - name of tab within the spreadsheet to be processes.
        """
        ws = self.wb[tabname]
        columns = self.get_column_positions(ws)
        start_row = ws["C3"].value
        end_row = ws["C4"].value

        for row in range(start_row, end_row+1):
            # Get skip value. Only care if str starting "y", so convert to
            # single lower-case character
            skip = ws.cell(row=row, column=columns["Skip"]).value
            skip = str(skip).lower()[:1]
            # Get action value,force to be str and remove any start/end spaces
            action = ws.cell(row=row, column=columns["Action"]).value
            action = str(action).strip()
            # Skip the row if skip has been set or no action has been set
            if action == "None" or skip == "y":
                continue
            #Extract arguments
            args = ws.cell(row=row, column=columns["Args"]).value
            if args:
                args = str(args)
                args = [e for e in args.split(",")]

            # Write execution time to spreadsheet
            now = time.strftime("(%H:%M:%S) %d/%m/%Y")
            ws.cell(row=row, column=columns["Runtime"]).value = now

            # Try to execute the action
            execution_failed = False
            try:
                if args:
                    result = getattr(actions, action)(*args)
                else:
                    result = getattr(actions, action)()
            except Exception as err:
                execution_failed = True
                result = "Error - row {}, action '{}': {}".format(row, action, str(err))

            # Write outcome to spreadsheet
            ws.cell(row=row, column=columns["Result"]).value = result

            # Result backgound colour highlighting
            if execution_failed:
                colour = "orange"
            # Apply conditional colour highlighting if condition set
            else:
                colour = ""# default empty colour
                
                condition = ws.cell(row=row, column=columns["Condition"]).value
                if condition:
                    colour = self.condition_check(result, condition)
            if colour:
               ws.cell(row=row, column=columns["Result"]).fill = self.fills[colour]
                
        #Save spreadsheet
        new_filename = self.save_results()
        print "Saved:", new_filename

    def condition_check(self, result, condition):
        """Apply conditional highlight check to result and return
        related background colour depending on the outcome.
        Args:
            result - the result value of a check
            condition - string containing conditional expresion that
            evaluates result in some way, e.g. "result == 5" or
            "'fail' not in result"
        Returns:
            colour name (needs to correspond to key in self.fills)
        """
        r = result # alias for convenience in condition
        try:
            check =  eval(condition)    
        except Exception as err:
            check = None
            print "Evaluation failed:", condition, err
            
        # Map outcomes to colours
        outcome_map = {True:"green", False: "red", None: "purple"}
        colour = outcome_map.get(check, None)       
        return colour

    def save_results(self):
        """Save results as new spreadsheet in results folder"""
        results_folder = os.path.join(os.getcwd(), "results")
        #Create the folder if it doesn't exist
        if not os.path.exists(results_folder):
            os.makedirs(results_folder)
        #Add yellow stripe to top row to make easier to distinguish for master Excel file
        for ws in self.wb.worksheets:
            for column in range(1, 18):
                ws.cell(row=1, column=column).fill = self.fills["yellow"]
        #Save the results
        result_filename = (os.path.splitext(self.filename)[0]
                           + time.strftime("_results_[%Y.%m.%d_%H.%M.%S].xlsx"))
        results_file = os.path.join(results_folder, result_filename)
        self.wb.save(results_file)
        return results_file

    def get_column_positions(self, ws, heading_row=6, max_column=20):
        """Create mapping of headings to column numbers and return as dict"""
        column_positions = {}
        for col in range(1, max_column+1):
            val = str(ws.cell(row=heading_row, column=col).value)
            if val:
                column_positions[val] = col
        return column_positions


if __name__ == "__main__":
    print "Started"
    filename = "py_runner.xlsx"
    go = Excel(filename)
    go.run_tab("Activities")
